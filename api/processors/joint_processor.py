"""
Joint Processor - Optimized for Vercel Serverless
Extracted from main2.py for better modularity and faster cold starts
"""

import pandas as pd
import os
import re
from typing import List, Dict, Tuple, Optional
import logging
from openpyxl import load_workbook
import time
import signal

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TimeoutError(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutError("Processing timeout")

class JointProcessor:
    def __init__(self, input_file: str, original_filename: str = None):
        self.input_file = input_file
        self.original_filename = original_filename
        self.price_records: List[Dict] = []
        self.type_records: List[Dict] = []
        self.price_id = 1
        self.type_id = 1
        self.description_map: Dict[str, str] = {}
        
        # Extract series name from filename
        self.series_name = self.extract_series_from_filename()
        
        # Cache for optimized reading
        self._wb = None
        self._sheets_cache = {}
    
    def extract_series_from_filename(self) -> str:
        """ดึงชื่อ series จากชื่อไฟล์ โดยจัดการกับ UUID และ timestamp"""
        if self.original_filename:
            base_name = os.path.splitext(self.original_filename)[0]
        else:
            base_name = os.path.splitext(os.path.basename(self.input_file))[0]
        
        # Clean up various patterns
        patterns_to_remove = [
            r'^\d{8}_\d{6}_[a-f0-9]{8}_',  # timestamp pattern
            r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_',  # UUID
            r'^[a-f0-9]{8}_',  # job_id pattern
        ]
        
        for pattern in patterns_to_remove:
            base_name = re.sub(pattern, '', base_name)
        
        # Remove unwanted suffixes/prefixes
        suffixes_to_remove = ['_data', '_price', '_export', '_backup', '_processed']
        prefixes_to_remove = ['data_', 'price_', 'export_', 'backup_', 'processed_']
        
        for suffix in suffixes_to_remove:
            if base_name.lower().endswith(suffix):
                base_name = base_name[:-len(suffix)]
                break
        
        for prefix in prefixes_to_remove:
            if base_name.lower().startswith(prefix):
                base_name = base_name[len(prefix):]
                break
        
        return base_name.strip().replace(' ', '_')
    
    def validate_file(self) -> bool:
        """Validate that the input file exists and is accessible"""
        if not os.path.exists(self.input_file):
            logger.error(f"ไม่เจอไฟล์ {self.input_file}")
            return False
        return True
    
    def get_optimized_workbook(self):
        """Get cached workbook with optimized settings"""
        if self._wb is None:
            self._wb = load_workbook(
                self.input_file, 
                read_only=True,
                data_only=True,
                keep_links=False
            )
        return self._wb
    
    def read_sheet_optimized(self, sheet_name_or_index, **kwargs):
        """Read sheet with optimized pandas settings"""
        cache_key = f"{sheet_name_or_index}_{str(kwargs)}"
        if cache_key not in self._sheets_cache:
            # Limit rows for performance on Vercel
            if 'nrows' not in kwargs:
                kwargs['nrows'] = 200  # Limit rows
                
            self._sheets_cache[cache_key] = pd.read_excel(
                self.input_file,
                sheet_name=sheet_name_or_index,
                engine='openpyxl',
                **kwargs
            )
        return self._sheets_cache[cache_key]
    
    def load_descriptions_from_sheet2(self) -> bool:
        """Load descriptions from sheet2 mapping Type to Description - OPTIMIZED"""
        try:
            logger.info("Loading descriptions from sheet2...")
            # Use optimized reading with limited rows
            df_sheet2 = self.read_sheet_optimized(1, dtype=str, nrows=100)
            
            # Strip whitespace from column names
            df_sheet2.columns = df_sheet2.columns.str.strip()
            
            # Find Type and Description columns
            type_col = None
            desc_col = None
            
            for i, col in enumerate(df_sheet2.columns):
                if 'type' in str(col).lower():
                    type_col = col
                    if i + 1 < len(df_sheet2.columns):
                        desc_col = df_sheet2.columns[i + 1]
                    break
            
            if type_col is None or desc_col is None:
                logger.warning("ไม่พบคอลัมน์ Type หรือ Description ใน sheet2")
                return False
            
            # Create mapping efficiently
            valid_mask = (df_sheet2[type_col].notna()) & (df_sheet2[type_col] != 'nan')
            valid_data = df_sheet2[valid_mask].head(50)  # Limit for performance
            
            for _, row in valid_data.iterrows():
                type_name = str(row[type_col]).strip()
                desc_text = str(row[desc_col]).strip() if pd.notna(row[desc_col]) else ''
                if type_name:
                    self.description_map[type_name] = desc_text
            
            logger.info(f"Loaded {len(self.description_map)} descriptions from sheet2")
            return True
            
        except Exception as e:
            logger.error(f"Error loading sheet2: {e}")
            return False
    
    def update_type_descriptions(self):
        """Update type records with descriptions from sheet2"""
        for record in self.type_records:
            type_name = record['Type']
            description = self.description_map.get(type_name, '')
            record['Description'] = description
    
    def clean_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize multi-level headers"""
        clean_cols = []
        for top, sub in df.columns:
            top_str = str(top).strip() if pd.notna(top) else ''
            sub_str = str(sub).strip() if pd.notna(sub) else ''
            clean_cols.append((top_str, sub_str))
        df.columns = pd.MultiIndex.from_tuples(clean_cols)
        return df
    
    def read_cell_background_color_optimized(self, sheet_name: str, row: int, col: int) -> str:
        """Read background color from Excel cell - OPTIMIZED"""
        try:
            wb = self.get_optimized_workbook()
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            cell = ws.cell(row=row + 1, column=col + 1)
            
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                color = str(cell.fill.start_color.rgb)
                if len(color) == 8 and color.startswith('FF'):
                    color = color[2:]
                
                if color == '00000000' or color == '000000' or not color:
                    return 'FFFFFF'
                
                return color
            return 'FFFFFF'
            
        except Exception as e:
            logger.warning(f"Cannot read cell color: {e}")
            return 'FFFFFF'
    
    def find_dimension_mode(self, sub_df: pd.DataFrame) -> Optional[str]:
        """Find the dimension mode (W first priority, then H)"""
        if 'W' in sub_df.columns:
            return 'W'
        elif 'H' in sub_df.columns:
            return 'H'
        return None
    
    def process_width_data(self, table_name: str, vals: pd.DataFrame, 
                          sheet_name: str = None) -> Tuple[float, float]:
        """Process width-based pricing data - OPTIMIZED"""
        w_vals = vals['W'].astype(float)
        p_vals = vals['Price'].astype(float)
        wmin, wmax = w_vals.min(), w_vals.max()
        
        price_col_idx = list(vals.columns).index('Price')
        
        # Limit processing for performance
        for idx, (w, p) in enumerate(list(zip(w_vals, p_vals))[:50]):  # Max 50 rows
            original_idx = vals.index[idx]
            
            color = 'FFFFFF'
            if sheet_name:
                color = self.read_cell_background_color_optimized(
                    sheet_name, original_idx + 2, price_col_idx
                )
            
            self.price_records.append({
                'ID': self.price_id,
                'Serie': self.series_name,
                'Type': table_name,
                'Width': w,
                'Height': 0,
                'Price': p,
                'Glass_QTY': 0,
                'Color': color
            })
            self.price_id += 1
        
        return wmin, wmax
    
    def process_height_data(self, table_name: str, vals: pd.DataFrame,
                           sheet_name: str = None) -> Tuple[float, float]:
        """Process height-based pricing data - OPTIMIZED"""
        h_vals = vals['H'].astype(float)
        p_vals = vals['Price'].astype(float)
        hmin, hmax = h_vals.min(), h_vals.max()
        
        price_col_idx = list(vals.columns).index('Price')
        
        # Limit processing for performance
        for idx, (h, p) in enumerate(list(zip(h_vals, p_vals))[:50]):  # Max 50 rows
            original_idx = vals.index[idx]
            
            color = 'FFFFFF'
            if sheet_name:
                color = self.read_cell_background_color_optimized(
                    sheet_name, original_idx + 2, price_col_idx
                )
            
            self.price_records.append({
                'ID': self.price_id,
                'Serie': self.series_name,
                'Type': table_name,
                'Width': 0,
                'Height': h,
                'Price': p,
                'Glass_QTY': 0,
                'Color': color
            })
            self.price_id += 1
        
        return hmin, hmax
    
    def add_type_record(self, table_name: str, wmin: float, wmax: float, 
                       hmin: float, hmax: float):
        """Add a type record with dimension ranges"""
        self.type_records.append({
            'ID': self.type_id,
            'Serie': self.series_name,
            'Type': table_name,
            'Description': '',  # Will be updated later
            'width_min': wmin,
            'width_max': wmax,
            'height_min': hmin,
            'height_max': hmax
        })
        self.type_id += 1
    
    def process_table(self, table_name: str, sub_df: pd.DataFrame, 
                     sheet_name: str = None) -> bool:
        """Process a single table from the Excel file - OPTIMIZED"""
        # Clean column names
        sub_df.columns = sub_df.columns.str.strip()
        
        # Find dimension mode
        mode = self.find_dimension_mode(sub_df)
        if mode is None:
            logger.warning(f"Skip {table_name}: no W or H column")
            return False
        
        # Check for Price column
        if 'Price' not in sub_df.columns:
            logger.warning(f"Skip {table_name}: no Price column")
            return False
        
        # Extract valid rows efficiently - limit for performance
        required_cols = [mode, 'Price']
        vals = sub_df[required_cols].dropna(how='any').head(50)  # Max 50 rows
        
        if vals.empty:
            logger.warning(f"Skip {table_name}: no valid {mode} + Price rows")
            return False
        
        try:
            # Process based on mode
            if mode == 'W':
                wmin, wmax = self.process_width_data(table_name, vals, sheet_name)
                hmin = hmax = 0
            else:  # mode == 'H'
                hmin, hmax = self.process_height_data(table_name, vals, sheet_name)
                wmin = wmax = 0
            
            # Add type record
            self.add_type_record(table_name, wmin, wmax, hmin, hmax)
            logger.info(f"Processed {table_name}: {len(vals)} rows")
            return True
            
        except Exception as e:
            logger.error(f"Error processing {table_name}: {e}")
            return False
    
    def save_results(self, job_id: str) -> None:
        """Save processed data to Excel files with simple names"""
        if self.price_records:
            price_filename = 'Price.xlsx'
            pd.DataFrame(self.price_records).to_excel(price_filename, index=False)
            logger.info(f"Saved {len(self.price_records)} price records to {price_filename}")
        
        if self.type_records:
            type_filename = 'Type.xlsx'
            pd.DataFrame(self.type_records).to_excel(type_filename, index=False)
            logger.info(f"Saved {len(self.type_records)} type records to {type_filename}")
    
    def process(self, job_id: str) -> bool:
        """Main processing function - OPTIMIZED for Vercel"""
        
        # Set timeout for Vercel (8 seconds)
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(8)
        
        try:
            if not self.validate_file():
                return False
            
            start_time = time.time()
            logger.info(f"Starting optimized processing of {self.input_file}")
            
            # Get optimized workbook for color reading
            wb = self.get_optimized_workbook()
            sheet_name = wb.sheetnames[0]  # First sheet name
            
            # Read main sheet with optimized settings and limited rows
            logger.info("Loading main sheet...")
            df = self.read_sheet_optimized(0, header=[0, 1], dtype=str, nrows=200)
            
            # Clean headers
            df = self.clean_headers(df)
            
            # Filter out empty top-level columns
            df = df.loc[:, df.columns.get_level_values(0) != '']
            
            # Process each table in order - limit for performance
            processed_count = 0
            table_names = list(df.columns.get_level_values(0).unique())[:5]  # Max 5 tables
            
            for table_name in table_names:
                # Check timeout
                if time.time() - start_time > 6:  # Leave 2 seconds for cleanup
                    break
                    
                if self.process_table(table_name, df[table_name].copy(), sheet_name):
                    processed_count += 1
            
            # Load descriptions from sheet2 - with timeout check
            if time.time() - start_time < 6:
                self.load_descriptions_from_sheet2()
                self.update_type_descriptions()
            
            # Save results
            self.save_results(job_id)
            
            logger.info(f"Processing complete: {processed_count} tables processed")
            return processed_count > 0
            
        except TimeoutError:
            logger.error("Processing timeout exceeded")
            raise TimeoutError("Processing timeout exceeded")
        except Exception as e:
            logger.error(f"Error during processing: {e}")
            return False
        finally:
            # Clean up resources
            signal.alarm(0)  # Cancel timeout
            if self._wb:
                self._wb.close()