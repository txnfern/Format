#!/usr/bin/env python3
"""
Matrix Processor - Optimized for Vercel Serverless
Extracted from main.py for better modularity and faster cold starts
"""

import os
import re
import math
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional
import time
import signal

class TimeoutError(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutError("Processing timeout")

class MatrixProcessor:
    def __init__(self, job_id: str):
        self.job_id = job_id
        
    def to_number(self, val):
        """Convert value to number, removing commas"""
        try:
            if val is None:
                return None
            
            str_val = str(val).strip()
            # Remove comma, space, and special characters
            clean_val = re.sub(r'[,\s]', '', str_val)
            clean_val = re.sub(r'[^\d.-]', '', clean_val)
            
            if not clean_val or clean_val in ['', '-', '.']:
                return None
                
            f = float(clean_val)
            if math.isnan(f):
                return None
            return int(f) if f.is_integer() else f
        except:
            return None

    def normalize_rgb(self, fill):
        """Convert ARGB color to RGB hex format - แก้ไขให้อ่านสีที่ถูกต้อง"""
        if not fill:
            return "FFFFFF"
        
        # ตรวจสอบ patternType ก่อน - เฉพาะ solid fill เท่านั้น
        if hasattr(fill, 'patternType') and fill.patternType:
            pattern_value = fill.patternType.value if hasattr(fill.patternType, 'value') else str(fill.patternType)
            # ถ้าไม่ใช่ solid pattern ให้ถือว่าไม่มีสี
            if pattern_value != 'solid':
                return "FFFFFF"
        else:
            # ถ้าไม่มี patternType ให้ถือว่าไม่มีสี
            return "FFFFFF"
        
        # รายการสีที่ไม่ต้องการ (Excel theme colors) - ไม่รวม 92CDDC
        excluded_colors = [
            "00000000",  # สีใส
            "F2F2F2"
        ]
        
        color_found = ""
        
        # Check fgColor
        if hasattr(fill, 'fgColor') and fill.fgColor:
            if hasattr(fill.fgColor, 'rgb') and fill.fgColor.rgb:
                color_str = str(fill.fgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        # Check bgColor
        if not color_found and hasattr(fill, 'bgColor') and fill.bgColor:
            if hasattr(fill.bgColor, 'rgb') and fill.bgColor.rgb:
                color_str = str(fill.bgColor.rgb).upper()
                if color_str == "00000000":
                    return "FFFFFF"
                elif len(color_str) == 8:
                    color_found = color_str[2:]
                elif len(color_str) == 6:
                    color_found = color_str
        
        # ตรวจสอบว่าเป็นสีที่ไม่ต้องการหรือไม่
        if color_found in excluded_colors:
            return "FFFFFF"
        
        return color_found if color_found else "FFFFFF"

    def find_thickness_matrix_in_column_a(self, ws, raw, thickness_num):
        """Find matrix with specific thickness label - หาจากคอลัมน์ A เท่านั้น"""
        thickness_patterns = [
            rf"Thk\.{thickness_num}",
            rf"\b{thickness_num}\b",
            rf"Thickness\s*{thickness_num}",
            rf"หนา\s*{thickness_num}",
            rf"ชั้น\s*{thickness_num}",
            rf"ระดับ\s*{thickness_num}"
        ]
        
        # หา thickness header ในคอลัมน์ A เท่านั้น (column index 0)
        for r in range(min(raw.shape[0], 50)):  # Limit search for performance
            if raw.shape[1] > 0:  # ตรวจสอบว่ามีคอลัมน์ A
                cell_val = str(raw.iat[r, 0]).strip() if raw.iat[r, 0] is not None else ""
                for pattern in thickness_patterns:
                    if re.search(pattern, cell_val, re.IGNORECASE):
                        return r
        
        return None

    def find_main_matrix(self, ws, raw):
        """Find main matrix (1 or h/w header) - หา 1 จากคอลัมน์ A, h/w จากทั่วไป"""
        # หาจาก 1 header ในคอลัมน์ A เท่านั้น
        for r in range(min(raw.shape[0], 50)):  # Limit search
            if raw.shape[1] > 0:  # ตรวจสอบว่ามีคอลัมน์ A
                cell_val = str(raw.iat[r, 0]).strip() if raw.iat[r, 0] is not None else ""
                # หา 1 header ในคอลัมน์ A
                if re.search(r"\b1\b", cell_val, re.IGNORECASE):
                    return r, 0  # ส่งคืน column = 0 (คอลัมน์ A)
        
        # ถ้าไม่พบ 1 header ให้หา h/w header แทน (ค้นหาทั่วไป - backward compatibility)
        for r in range(min(raw.shape[0], 50)):
            for c in range(min(raw.shape[1], 20)):
                if raw.iat[r, c] is None:
                    continue
                if isinstance(raw.iat[r, c], str):
                    if re.search(r"\bh\s*/\s*w\b", raw.iat[r, c], re.IGNORECASE):
                        return r, c
        
        return None, None

    def process_file(self, input_file: str, output_dir: str, original_filename: str = None):
        """Process the Excel file with timeout protection"""
        
        # Set timeout for Vercel (8 seconds)
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(8)
        
        try:
            start_time = time.time()
            
            if original_filename:
                base_name = os.path.splitext(original_filename)[0]
            else:
                base_name = os.path.splitext(os.path.basename(input_file))[0]
                # ลบ UUID ออกจากชื่อไฟล์ (UUID format: 8-4-4-4-12 characters)
                uuid_pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_'
                base_name = re.sub(uuid_pattern, '', base_name)
            
            # Load Excel with read-only optimization
            xls = pd.ExcelFile(input_file, engine="openpyxl")
            wb = load_workbook(input_file, data_only=True, read_only=True)
            
            # Quick scan - limit to first 5 sheets for performance
            max_matrices_count = 1
            all_sheet_matrices = {}
            
            sheet_count = 0
            for sheet_name in xls.sheet_names[:5]:  # Limit sheets for Vercel
                if sheet_name.strip().lower() == "สารบัญ":
                    continue
                    
                sheet_count += 1
                if sheet_count > 3:  # Limit to 3 sheets for timeout
                    break
                
                try:
                    # Quick read with limited rows
                    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, 
                                      engine="openpyxl", nrows=100)
                    ws = wb[sheet_name]
                    
                    # หา main matrix
                    hr, hc = self.find_main_matrix(ws, raw)
                    if hr is None:
                        all_sheet_matrices[sheet_name] = []
                        continue
                    
                    # หา matrices - limit search for performance
                    found_matrices = [1]
                    for thickness in range(2, 6):  # Reduced from 20 to 6
                        hr_thick = self.find_thickness_matrix_in_column_a(ws, raw, thickness)
                        if hr_thick is not None:
                            found_matrices.append(thickness)
                        else:
                            break
                    
                    all_sheet_matrices[sheet_name] = found_matrices
                    if len(found_matrices) > max_matrices_count:
                        max_matrices_count = len(found_matrices)
                        
                except Exception as e:
                    all_sheet_matrices[sheet_name] = []
            
            # สร้าง template คอลัมน์
            matrix_columns = []
            for i in range(1, max_matrices_count + 1):
                matrix_columns.append(f"{i}_Color")
            
            price_rows = []
            type_rows = []
            price_id = 1
            type_id = 1
            processed_sheets = 0
            skipped_sheets = []
            
            for sheet in list(all_sheet_matrices.keys())[:3]:  # Process max 3 sheets
                # Check timeout
                if time.time() - start_time > 6:  # Leave 2 seconds for cleanup
                    break
                    
                available_matrices = all_sheet_matrices.get(sheet, [])
                if not available_matrices:
                    skipped_sheets.append({"sheet": sheet, "reason": "ไม่พบ matrix ใดๆ"})
                    continue
                
                try:
                    # Quick read with limited data
                    raw = pd.read_excel(xls, sheet_name=sheet, header=None, 
                                      engine="openpyxl", nrows=200)
                    ws = wb[sheet]
                    
                    # Find Glass_QTY and Description - limited search
                    sheet_glass_qty = 1
                    sheet_description = ""
                    
                    for r in range(min(raw.shape[0], 50)):
                        for c in range(min(raw.shape[1] - 1, 10)):
                            if raw.iat[r, c] is None:
                                continue
                            cell = str(raw.iat[r, c]).strip()
                            low = cell.lower()
                            
                            if low in ("glass_qty", "glass qty"):
                                next_cell = raw.iat[r, c + 1]
                                qty = self.to_number(next_cell)
                                if qty is not None:
                                    sheet_glass_qty = qty
                                
                            elif low == "description":
                                desc = raw.iat[r, c + 1]
                                if desc is not None:
                                    sheet_description = str(desc).strip()
                    
                    # Find main matrix
                    hr, hc = self.find_main_matrix(ws, raw)
                    if hr is None or hc is None:
                        skipped_sheets.append({"sheet": sheet, "reason": "ไม่พบ main matrix"})
                        continue
                    
                    # Read dimensions - limited for performance
                    widths = []
                    for c in range(hc + 1, min(raw.shape[1], hc + 21)):  # Max 20 widths
                        v = self.to_number(raw.iat[hr, c])
                        if v is None:
                            break
                        widths.append(v)
                    
                    heights = []
                    for r in range(hr + 1, min(raw.shape[0], hr + 21)):  # Max 20 heights
                        h_val = self.to_number(raw.iat[r, hc])
                        if h_val is None:
                            break
                        heights.append(h_val)
                    
                    if not widths or not heights:
                        skipped_sheets.append({"sheet": sheet, "reason": "ไม่พบ dimensions"})
                        continue
                    
                    # อ่านสีจาก matrices - optimized
                    matrix_colors = {}
                    
                    # อ่าน matrix 1 (main matrix)
                    if 1 in available_matrices:
                        matrix_colors[1] = self.read_color_matrix_fast(ws, raw, hr, hc, widths, heights)
                    
                    # อ่าน matrices อื่นๆ - limited
                    for thickness in available_matrices[:3]:  # Max 3 thickness matrices
                        if thickness == 1:
                            continue
                        
                        hr_thick = self.find_thickness_matrix_in_column_a(ws, raw, thickness)
                        if hr_thick is not None:
                            colors = self.read_color_matrix_with_thickness_row_fast(
                                ws, raw, hr, hc, hr_thick, widths[:10], heights[:10]  # Limited
                            )
                            matrix_colors[thickness] = colors
                    
                    # Create records
                    type_rows.append({
                        "ID": type_id,
                        "Serie": base_name,
                        "Type": sheet.strip(),
                        "Description": sheet_description,
                        "width_min": min(widths),
                        "width_max": max(widths),
                        "height_min": min(heights),
                        "height_max": max(heights),
                    })
                    type_id += 1
                    
                    # Create Price records - limited for performance
                    for i_h, h in enumerate(heights[:10]):  # Max 10 heights
                        for i_w, w in enumerate(widths[:10]):  # Max 10 widths
                            raw_price = raw.iat[hr + 1 + i_h, hc + 1 + i_w]
                            p = self.to_number(raw_price)
                            if p is None:
                                continue
                            
                            price_record = {
                                "ID": price_id,
                                "Serie": base_name,
                                "Type": sheet.strip(),
                                "Width": w,
                                "Height": h,
                                "Price": p,
                                "Glass_QTY": sheet_glass_qty,
                            }
                            
                            # เพิ่มคอลัมน์สี
                            for i in range(1, max_matrices_count + 1):
                                color_key = f"{i}_Color"
                                if i in matrix_colors:
                                    color_value = matrix_colors[i].get((h, w), "FFFFFF")
                                else:
                                    color_value = "FFFFFF"
                                price_record[color_key] = color_value
                            
                            price_rows.append(price_record)
                            price_id += 1
                    
                    processed_sheets += 1
                    
                except Exception as e:
                    skipped_sheets.append({"sheet": sheet, "reason": f"Error: {str(e)}"})
            
            # Save output files
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)
            
            price_file = output_path / f"Price_{self.job_id}.xlsx"
            type_file = output_path / f"Type_{self.job_id}.xlsx"
            
            pd.DataFrame(price_rows).to_excel(price_file, index=False)
            pd.DataFrame(type_rows).to_excel(type_file, index=False)
            
            return {
                "price_file": str(price_file),
                "type_file": str(type_file),
                "total_records": len(price_rows),
                "processed_sheets": processed_sheets,
                "skipped_sheets": skipped_sheets,
                "warnings": []
            }
            
        except TimeoutError:
            raise TimeoutError("Processing timeout exceeded")
        except Exception as e:
            raise Exception(f"Processing failed: {str(e)}")
        finally:
            signal.alarm(0)  # Cancel timeout
    
    def read_color_matrix_fast(self, ws, raw, hr, hc, widths, heights):
        """Fast color reading for main matrix"""
        color_map = {}
        for i_h, h in enumerate(heights[:10]):  # Limit for performance
            for i_w, w in enumerate(widths[:10]):
                try:
                    excel_row = hr + 2 + i_h
                    excel_col = hc + 2 + i_w
                    
                    if excel_row <= ws.max_row and excel_col <= ws.max_column:
                        cell = ws.cell(row=excel_row, column=excel_col)
                        color = self.normalize_rgb(cell.fill)
                        color_map[(h, w)] = color
                    else:
                        color_map[(h, w)] = "FFFFFF"
                except Exception:
                    color_map[(h, w)] = "FFFFFF"
        return color_map
    
    def read_color_matrix_with_thickness_row_fast(self, ws, raw, hr_main, hc_main, hr_thick, widths, heights):
        """Fast color reading for thickness matrices"""
        colors = {}
        # Use simple offset without optimization loop for speed
        row_offset, col_offset = 1, 1
        
        for i_h, h in enumerate(heights):
            for i_w, w in enumerate(widths):
                try:
                    excel_row = hr_thick + row_offset + i_h
                    excel_col = hc_main + col_offset + i_w
                    
                    if excel_row <= ws.max_row and excel_col <= ws.max_column:
                        cell = ws.cell(row=excel_row, column=excel_col)
                        color = self.normalize_rgb(cell.fill)
                        colors[(h, w)] = color
                    else:
                        colors[(h, w)] = "FFFFFF"
                except:
                    colors[(h, w)] = "FFFFFF"
        
        return colors